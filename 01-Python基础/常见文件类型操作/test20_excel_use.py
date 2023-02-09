# encoding=utf-8

"""
Excel格式文件操作:读取、写入
author：LBL
date:2023-2-9
"""
import openpyxl  # python使用openpyxl类操作excel文件


def create_excel(f_name):
    """ 创建excel文件(xls,xlsx)"""
    wb = openpyxl.Workbook()                # 创建空白的excel文件
    ws = wb["Sheet"]                	    # 设置初始工作表为Sheet表
    ws.title = 'my_sheet'           	    # 更改工作表表名
    wb.create_sheet()               	    # 创建新工作表，默认名为SheetN ,N为数字
    wb.create_sheet(index=2,title='my2')    # 创建新工作表，指定下标与名称。表按下标排序
    wb.save(f_name)                         # 保存excel文件，指定文件名
    wb.close()                      	    # 关闭文件对象，需要？


def write_excel():
    """ 写入单元格内容，要求文件没有被其他软件打开 """
    f_name = "test.xlsx"
    wb = openpyxl.load_workbook(f_name)         # 创建excel文件对象，参数为文件名
    ws = wb['Sheet']                        	# 创建表对象，指定表名
    ws.cell(1,5).value = "test_data"        	# 指定（行，列）设置其value属性，
    ws['A2'].value = "test"                 	# 指定单元格坐标，设置值
    wb.save(f_name)                         	# 保存
    wb.close()                              	# 关闭文件对象


def read_excel1(f_name):
    """ 读取指定范围内的单元格内容 """
    wb = openpyxl.load_workbook(f_name)     # 创建excel文件对象，参数为文件名
    ws = wb['Sheet']                       	# 创建表对象，指定表名
    for data in ws["A1":"E3"]:              # 读取A1到E3区域内的单元格内容，None也会读取
        for cell in data:                   # ws是表对象，data是指定范围内的一行单元格组成的元组，cell是具体单元格对象
            print(cell.value,end = " ")
        print()


def read_excel2(f_name):
    """ 读取某些列不知道行数的全部表格内容 """
    wb = openpyxl.load_workbook(f_name)         # 创建excel文件对象，参数为文件名
    ws = wb["Sheet"]                        	# 创建表对象，指定表名
    first_name = ["A","E"]                  	# 指定要读取的列名称
    last_name = [str(x) for x in range(1,ws.max_row+1)]     # ws.max_row为文件中最大行数
    list1 = []
    for i in first_name:                        # 循环将A1，A2，A3..E1，E2..的值添加到列表中
        for j in last_name:
            list1.append(ws[i+j].value)
    print(list1)


# 程序主入口
if __name__ == "__main__":
    # excel文件分xls与xlsx,一个是二进制存储方式一个是xml存储方式
    # 文件是一个有序序列，每个元素是一个表，
    # 每个表是一个有序序列，按下标取单元格
    # 获取最大列数.max_column    获取最大行数 .max_row    读取单元格内容.value
    create_excel("test.xlsx")
    write_excel()
    read_excel1("test.xlsx")
    read_excel2("test.xlsx")





